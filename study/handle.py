from openpyxl import Workbook, load_workbook

import operator

# 初航舰 xls 文件对象
from study.baobiao_parse import parse
from utils.excel.gen_head import gen_head

year = 2021
# excel 表名称
excel_name = "商户费用应收表--2021.1.19--2021 更新于2021.1.20-test-01"

# load workbook
wb = load_workbook('../data/商户费用应收表--2021.1.19--2021 更新于2021.1.20.xlsx')
wb.template = False
ws = wb.active
target = wb.copy_worksheet(ws)
# sheets
print(wb.sheetnames)
# sheetLen = len(sheets)
for sheet in wb:
    print('sheet:', sheet, ', type: ', type(sheet))
    # 生成表头
    gen_head(wb, year)
    # 填充内容
    # row: cellIndex
    for n in range(sheet.max_row):
        row: int = n + 2
        cell_index: int = 9
        cell = ws.cell(row=row, column=cell_index + 1)
        data_map: dict = parse(cell.value, year)
        print('data_map:', data_map)
        # 获取 2021 年数据
        if data_map is None:
            continue
        # 填充到列
        month = data_map['month']
        for m in range(month):
            cell_index += 2
            # key
            key = str(year) + '-' + str(m + 1)
            # print('key： ', key, ', data map', data_map)
            data = data_map[key]
            # 租金
            # print('data:', data)
            ws.cell(row=row, column=cell_index, value=str(data))
        break
    break


print('保存 excel')
wb.save('../output/test-01.xls')

