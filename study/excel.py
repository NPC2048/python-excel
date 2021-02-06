from openpyxl import Workbook, load_workbook


wb = load_workbook('./data/商户费用应收表--2021.1.19--2021 更新于2021.1.20-test-01.xlsx')
wb.template = False
ws = wb.active

print(wb.sheetnames)

year = 2021
cellIndex = 9
for i in range(12):
    title1 = str(year) + '''年''' + str(i) + '月'
    title2 = str(i+1) + '月提成租金'
    ws.cell(row=1, column=cellIndex, value=title1)
    ws.cell(row=1, column=cellIndex+1, value=title2)
    cellIndex += 2
ws.cell(row=1, column=cellIndex, value='租金合计')
wb.save('./output/new.xlsx')
wb.close()