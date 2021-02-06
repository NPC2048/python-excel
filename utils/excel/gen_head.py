# 生成表头
from openpyxl import Workbook


def gen_head(workbook: Workbook, title_year: int):
    cellIndex: int = 9
    work_sheet = workbook.active
    for i in range(12):
        title1 = str(title_year) + '''年''' + str(i) + '月'
        title2 = str(i + 1) + '月提成租金'
        work_sheet.cell(row=1, column=cellIndex, value=title1)
        work_sheet.cell(row=1, column=cellIndex + 1, value=title2)
        cellIndex += 2
    work_sheet.cell(row=1, column=cellIndex, value='租金合计')

