# 如果没有就先安装
# cmd: pip3 install openpyxl
# 导入 openpyxl
from openpyxl import Workbook, load_workbook
import operator


def main():
    # 读取 excel
    wb = load_workbook(filename='./data/2021年1月收款明细2021.2.2.xlsx', read_only=True)
    wb.template = False
    # 字典
    data_dict: dict = {}

    # 具体操作查 openpyxl
    for sheet in wb:
        print(sheet.title)
        # 循环查看 sheet 的每一行
        begin_index = 3
        for i in range(sheet.max_row):
            # 如果
            index = i + begin_index
            # 序号
            no: str = sheet.cell(row=index, column=1).value
            # 如果为空，说明到底了，结束循环
            if no is None or operator.eq(str(no), '') or str(no).find('合计') > -1:
                break
            # 字典 key
            data = read_data(sheet, index)
            key = str(data.get('store_name')) + '-' + str(data.get('brand'))
            # 将数据以 key-value 的形式放入字典
            # 判断字段是否存在
            arr: list = data_dict.get(key)

            if arr is None:
                data_dict[key] = [data]
            else:
                arr.append(data)

    merge_data(data_dict)
    print(data_dict)


# 从 excel 读取数据
def read_data(sheet, index):
    # 序号
    no = sheet.cell(row=index, column=2).value
    # 收款日期
    rec_date = sheet.cell(row=index, column=2).value
    # 收据号码
    rec_no = sheet.cell(row=index, column=3).value
    # 合同号
    contact_no = sheet.cell(row=index, column=4).value
    # 商铺编号
    store_no = sheet.cell(row=index, column=5).value
    # 商户姓名
    store_name = sheet.cell(row=index, column=6).value
    # 品牌
    brand = sheet.cell(row=index, column=7).value
    # 意向金
    intention_money = sheet.cell(row=index, column=8).value
    # pos 机押金
    pos_money = sheet.cell(row=index, column=9).value
    # 履约保证金
    promise_money = sheet.cell(row=index, column=10).value
    # 租金
    rent = sheet.cell(row=index, column=11).value
    # 物业管理费
    property_manage = sheet.cell(row=index, column=12).value
    # 推广费
    advert = sheet.cell(row=index, column=13).value
    return {
        'no': no,
        'rec_date': rec_date,
        'rec_no': rec_no,
        'contact_no': contact_no,
        'store_no': store_no,
        'store_name': store_name,
        'brand': brand,
        'intention_money': intention_money,
        'pos_money': pos_money,
        'promise_money': promise_money,
        'rent': rent,
        'property_manage': property_manage,
        # 推廣費
        'advert': advert,
        # 水費
        'water_charge': sheet.cell(row=index, column=14).value,
        # 装修保证金
        'decoration_deposit': sheet.cell(row=index, column=15).value,
        # 装修管理费
        'decoration_management_fee': sheet.cell(row=index, column=16).value,
        # 垃圾清运费
        'garbage_clean_fee': sheet.cell(row=index, column=17).value,
        # 围挡费
        'enclosure_fee': sheet.cell(row=index, column=18).value,
        # 装修期电费
        'decoration_period_electricity_fee': sheet.cell(row=index, column=19).value,
        # 电费
        'electricity_fee': sheet.cell(row=index, column=20).value,
        # 广告灯箱费
        'advert_light': sheet.cell(row=index, column=21).value,
        # 广告位费用
        'advert_space_fee': sheet.cell(row=index, column=22).value,
        # 应收合计
        'total_receivables': sheet.cell(row=index, column=15).value,
        # 实收合计
        'total_paid': sheet.cell(row=index, column=16).value
    }


# 合并数据
def merge_data(data_dict: dict):
    for key in data_dict:
        print('key: ', key)


main()